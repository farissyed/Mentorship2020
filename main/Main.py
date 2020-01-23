import numpy as np
from openpyxl import load_workbook
from main import Country
from main import Reader

class Game:
    end_game = False
    print('Welcome to the the Battle Simulator')
    #Add country reader
    # countries = np.zeros((10, 10))
    # str = input("Enter country one\n")
    # wb = load_workbook('CountryData.xlsx')  # opens the Excel file
    # source = wb.active
    # ws = wb.copy_worksheet(source)
    # c1 = Country
    # c2 = Country
    #
    # for i in range(2, 11):
    #     country = ws.cell(row=i, column=1)
    #     if (country.value == str):
    #         c1.ID_TAG = ws.cell(row=i, column=2).value
    #         c1.GDP = ws.cell(row=i, column=3).value
    #         c1.ARMY_ASSETS = ws.cell(row=i, column=4).value
    #         c1.NAVY_ASSETS = ws.cell(row=i, column=5).value
    #         c1.AIR_FORCE_ASSETS = ws.cell(row=i, column=6).value
    #         c1.MANPOWER = ws.cell(row=i, column=7).value
    #         c1.NUCLEAR_CAPABILITIES = ws.cell(row=i, column=8).value
    #
    # str = input("Enter country two\n")
    # for i in range(2, 11):
    #     country = ws.cell(row=i, column=1)
    #     if (country.value == str):
    #         c2.ID_TAG = ws.cell(row=i, column=2).value
    #         c2.GDP = ws.cell(row=i, column=3).value
    #         c2.ARMY_ASSETS = ws.cell(row=i, column=4).value
    #         c2.NAVY_ASSETS = ws.cell(row=i, column=5).value
    #         c2.AIR_FORCE_ASSETS = ws.cell(row=i, column=6).value
    #         c2.MANPOWER = ws.cell(row=i, column=7).value
    #         c2.NUCLEAR_CAPABILITIES = ws.cell(row=i, column=8).value
    #Add maneuver matrix
    maneuvers = np.array(144, dtype='string')
    wb1 = load_workbook('MilitaryTactics.xlsx')  # opens the Excel file
    source1 = wb1.active
    ws = wb1.copy_worksheet(source1)
    print(ws.cell(row=1,column=1).value)
    for i in range(2,12):
        for j in range(2,12):
            print('not done')
    moves = np.zeros((10,10))           #initialize moves
    while(end_game == False):
        military_type = input("Entering the following number for:\n[0] Offense\n[1] Defense\n")
        if(military_type == '0'):
            maneuver = input("Pick one of the following offensive maneuvers: \n[0] Push Middle\n[1] Flank Right\n[2] Flank Left\n[3] Pincer\n[4] Push Equally\n[5]Trap on Two Fronts\n[999] Go back to beginning\n")
            if(maneuver == '0'):
                print('You tried to push middle\n')
            elif (maneuver == '1'):
                print('You tried to flank right\n')
            elif (maneuver == '2'):
                print('You tried to flank left\n')
            elif (maneuver == '3'):
                print('You tried to do a pincer movement\n')
            elif (maneuver == '4'):
                print('You tried to push equally on all fronts\n')
            elif (maneuver == '5'):
                print('You tried to trap the enemy on two fronts\n')
            elif(maneuver == '999'):
                break
            else:
                print('Wrong input. Try again.')
        if(military_type == '1'):
            maneuver = input('Pick one of the following defensive maneuvers:\n[0] Retreat\n[1] Retreating Fire\n[2] Find and Hold a Choke Point\n[3] Turtle')
            if (maneuver == '0'):
                print('You tried to retreat\n')
            elif (maneuver == '1'):
                print('You tried to retreat to cover fire\n')
            elif (maneuver == '2'):
                print('You tried to find and hold a choke point\n')
            elif (maneuver == '3'):
                print('You tried to create a defensive turtle for a better defense\n')
            elif (maneuver == '999'):
                break
            else:
                print('Wrong input. Try again.')
        print(moves[int(maneuver)][int(military_type)])